import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

def create_readiness_report(output_path):
    wb = Workbook()
    
    # ----------------------------------------------------
    # Color Palette Definitions (Teal Theme matching the App)
    # ----------------------------------------------------
    COLOR_TEAL_DARK = "005F5F"     # Header
    COLOR_TEAL_MEDIUM = "008B8B"   # Section Header
    COLOR_TEAL_LIGHT = "E0F2F1"    # Highlight Soft Teal
    COLOR_WHITE = "FFFFFF"
    COLOR_GRAY_LIGHT = "F9F9F9"    # Alternating row
    COLOR_GRAY_BORDER = "D3D3D3"
    
    # Status Colors
    COLOR_CRITICAL_FILL = "FADBD8" # Soft Red
    COLOR_CRITICAL_TEXT = "78281F"
    COLOR_MEDIUM_FILL = "FCF3CF"   # Soft Yellow
    COLOR_MEDIUM_TEXT = "7E5109"
    COLOR_LOW_FILL = "D4EFDF"      # Soft Green
    COLOR_LOW_TEXT = "196F3D"
    COLOR_ACTION_FILL = "F5CBA7"   # Soft Orange
    COLOR_ACTION_TEXT = "7E5109"
    
    # Fonts
    font_title = Font(name="Calibri", size=18, bold=True, color="000000")
    font_section = Font(name="Calibri", size=13, bold=True, color="000000")
    font_header = Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE)
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=11)
    
    # Fills
    fill_header = PatternFill(start_color=COLOR_TEAL_DARK, end_color=COLOR_TEAL_DARK, fill_type="solid")
    fill_section = PatternFill(start_color=COLOR_TEAL_LIGHT, end_color=COLOR_TEAL_LIGHT, fill_type="solid")
    fill_zebra = PatternFill(start_color=COLOR_GRAY_LIGHT, end_color=COLOR_GRAY_LIGHT, fill_type="solid")
    
    # Status/Severity Fills & Fonts
    fill_critical = PatternFill(start_color=COLOR_CRITICAL_FILL, end_color=COLOR_CRITICAL_FILL, fill_type="solid")
    font_critical = Font(name="Calibri", size=11, bold=True, color=COLOR_CRITICAL_TEXT)
    
    fill_medium = PatternFill(start_color=COLOR_MEDIUM_FILL, end_color=COLOR_MEDIUM_FILL, fill_type="solid")
    font_medium = Font(name="Calibri", size=11, bold=True, color=COLOR_MEDIUM_TEXT)
    
    fill_low = PatternFill(start_color=COLOR_LOW_FILL, end_color=COLOR_LOW_FILL, fill_type="solid")
    font_low = Font(name="Calibri", size=11, bold=True, color=COLOR_LOW_TEXT)
    
    fill_action = PatternFill(start_color=COLOR_ACTION_FILL, end_color=COLOR_ACTION_FILL, fill_type="solid")
    font_action = Font(name="Calibri", size=11, bold=True, color=COLOR_ACTION_TEXT)

    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Borders
    thin_border = Border(
        left=Side(style='thin', color=COLOR_GRAY_BORDER),
        right=Side(style='thin', color=COLOR_GRAY_BORDER),
        top=Side(style='thin', color=COLOR_GRAY_BORDER),
        bottom=Side(style='thin', color=COLOR_GRAY_BORDER)
    )

    # ====================================================
    # TAB 1: SUMMARY DASHBOARD
    # ====================================================
    ws_dash = wb.active
    ws_dash.title = "Summary Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Sheet Title
    ws_dash["B2"] = "Oral Health AI - Production Readiness Assessment"
    ws_dash["B2"].font = font_title
    
    # Description
    ws_dash["B3"] = "This workbook evaluates the core FastAPI backend and Flutter client against standard production guidelines."
    ws_dash["B3"].font = Font(name="Calibri", size=11, italic=True)

    # Summary KPI block
    ws_dash["B5"] = "Assessment Overview"
    ws_dash["B5"].font = font_section
    
    headers_dash = ["METRIC", "COUNT"]
    for col_idx, header in enumerate(headers_dash, start=2):
        cell = ws_dash.cell(row=6, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        
    metrics = [
        ("Total Checklist Areas Evaluated", 10),
        ("Critical Action Items Required", 0),
        ("Medium Action Items Required", 0),
        ("Low Action Items Required", 0),
        ("Ready for Production", 10)
    ]
    
    for row_idx, (metric, val) in enumerate(metrics, start=7):
        # Label
        cell_label = ws_dash.cell(row=row_idx, column=2, value=metric)
        cell_label.font = font_bold
        cell_label.border = thin_border
        cell_label.alignment = align_left
        
        # Value
        cell_val = ws_dash.cell(row=row_idx, column=3, value=val)
        cell_val.font = font_regular
        cell_val.alignment = align_center
        cell_val.border = thin_border
        
        # Color codes
        if "Critical" in metric:
            cell_val.fill = fill_critical if val > 0 else PatternFill(fill_type=None)
            cell_val.font = font_critical if val > 0 else font_regular
        elif "Medium" in metric:
            cell_val.fill = fill_medium if val > 0 else PatternFill(fill_type=None)
            cell_val.font = font_medium if val > 0 else font_regular
        elif "Low" in metric:
            cell_val.fill = fill_low if val > 0 else PatternFill(fill_type=None)
            cell_val.font = font_low if val > 0 else font_regular
        elif "Ready" in metric:
            if val == 10:
                cell_val.fill = fill_low
                cell_val.font = font_low
            else:
                cell_val.fill = fill_critical
                cell_val.font = font_critical
            
    # Add Pie Chart for Action Items
    chart = PieChart()
    labels = Reference(ws_dash, min_col=2, min_row=8, max_row=10)
    data = Reference(ws_dash, min_col=3, min_row=7, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title = "Outstanding Action Items by Severity"
    ws_dash.add_chart(chart, "E5")
    
    # Autofit Summary Columns
    ws_dash.column_dimensions["B"].width = 35
    ws_dash.column_dimensions["C"].width = 12
    ws_dash.column_dimensions["D"].width = 5
    
    # ====================================================
    # TAB 2: DETAILED FINDINGS
    # ====================================================
    ws_details = wb.create_sheet("Assessment Details")
    ws_details.views.sheetView[0].showGridLines = True
    
    ws_details["A2"] = "Detailed Production Readiness Findings & Improvements"
    ws_details["A2"].font = font_title
    
    headers_details = [
        "ID", "Area Checked", "Current Status", 
        "Severity", "Suggested Improvement", "File Name"
    ]
    
    for col_idx, header in enumerate(headers_details, start=1):
        cell = ws_details.cell(row=4, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
        ws_details.row_dimensions[4].height = 25
        
    findings = [
        (
            "R-01", "JWT Authentication Flow", "Resolved", "Critical",
            "1. User Active Verification: Implemented verification of user existence and active status in DB within get_current_user() decorator.\n2. Enforce SECRET_KEY: Secret key fallback removed; runtime exception raised at start if SECRET_KEY env is missing.\n3. Secure Client Storage: Integrated flutter_secure_storage to encrypt JWTs securely.\n4. Intercept 401s: Configured ApiService to catch 401 errors, clear tokens, and route to login.",
            "backend/app/utils/oauth2.py\nbackend/app/config.py\nlib/services/api_service.dart"
        ),
        (
            "R-02", "Password Hashing", "Resolved", "Medium",
            "1. Enforce Complexity: Added min 8 chars length validation in Pydantic UserSignup schema.\n2. Secure Reset Route: Replaced untyped dict reset-password body with validated ResetPasswordRequest Pydantic model.",
            "backend/app/schemas/user_schema.py\nbackend/app/routes/auth_routes.py"
        ),
        (
            "R-03", "Environment Variables", "Resolved", "Critical",
            "1. Standard Cloud Loading: Made load_dotenv() conditional to prioritize container host overrides.\n2. Prevent Git Leaks: Updated root .gitignore to exclude *.env, .env, and venv/ folders.\n3. Dynamic Frontend URLs: Set up String.fromEnvironment for baseUrl configuration in Flutter client.",
            "backend/app/config.py\n.gitignore\nlib/services/api_service.dart"
        ),
        (
            "R-04", "CORS Configuration", "Resolved", "Critical",
            "1. Restrict Wildcard: Restricted origins dynamically in main.py, removing wildcard allow_origins with credentials.\n2. Dynamic Loading: Configured main.py to load allowed origins from environment variables.",
            "backend/app/main.py"
        ),
        (
            "R-05", "File Upload Checks", "Resolved", "Critical",
            "1. Validate MIME/Magic Bytes: Added validation of file headers/magic bytes for png/jpeg/webp formats.\n2. Max Upload Sizes: Limited upload sizes to 5MB before stream reads to prevent memory exhaustion.\n3. Client Checks: Added local file size check in ApiService before dispatching uploads.",
            "backend/app/routes/image_routes.py\nlib/services/api_service.dart"
        ),
        (
            "R-06", "API Error Handling", "Resolved", "Medium",
            "1. Handle DB Exceptions: Wrapped database commit operations in transactions with rollback and proper HTTP error responses.\n2. Async Email Delivery: Offloaded email sending to BackgroundTasks and wrapped smtplib in try-except.\n3. Dynamic Client Exceptions: Implemented proper network timeout limits (10s) and structured exception raising in Flutter ApiService.",
            "backend/app/routes/auth_routes.py\nbackend/app/utils/email_service.py\nlib/services/api_service.dart"
        ),
        (
            "R-07", "DB Connection Handling", "Resolved", "Critical",
            "1. Connect Pool Parameters: Optimized database engine with pool_size, max_overflow, pool_recycle, and pool_pre_ping settings.\n2. Schema Migration Tool: Disabled synchronous create_all schema creation in production env.\n3. Foreign Key Indexes: Added index=True to user_id foreign keys in Scan, DailyLog, and Reminder models.",
            "backend/app/database.py\nbackend/app/main.py\nbackend/app/models/scan_model.py"
        ),
        (
            "R-08", "Pydantic Input Validation", "Resolved", "Medium",
            "1. Standardize Request Models: Added strict Pydantic models for ResetPasswordRequest and UpdateProfileRequest.\n2. Consolidate Schemas: Moved inline route requests to centralized schemas folder.\n3. Numerical Limits: Added ge=0, le=10 limits to symptom scales in DailyLog Pydantic schemas.",
            "backend/app/routes/auth_routes.py\nbackend/app/routes/daily_log_routes.py\nbackend/app/routes/reminder_routes.py"
        ),
        (
            "R-09", "Logging Configuration", "Resolved", "Low",
            "1. Structured Logger: Configured Python logging standard module with proper console handlers.\n2. Purge Print Statements: Substituted backend print() calls with appropriate logger statements.\n3. Strip Release Logs: Silenced Flutter print output in release builds using kReleaseMode validation.",
            "backend/app/main.py\nbackend/app/utils/ai_service.py\nlib/main.dart"
        ),
        (
            "R-10", "Deployment Readiness", "Resolved", "Critical",
            "1. Non-blocking Async I/O: Offloaded blocking image vision model requests to a separate worker thread pool via anyio.to_thread.run_sync.\n2. Dynamic CDN/S3 Uploads: Refactored image output URLs to load domains from config.\n3. Timezone Agnostic: Updated Flutter client to dynamically resolve local timezone at boot using flutter_timezone.\n4. Global Client Monitors: Containerized backend with Gunicorn and a multi-stage production Dockerfile.",
            "backend/app/routes/image_routes.py\nlib/main.dart\nbackend/app/utils/ai_service.py"
        )
    ]
    
    for row_idx, f_data in enumerate(findings, start=5):
        id_val, area, status, severity, improvement, files = f_data
        
        ws_details.cell(row=row_idx, column=1, value=id_val).alignment = align_center
        ws_details.cell(row=row_idx, column=2, value=area).alignment = align_left
        
        # Status
        cell_status = ws_details.cell(row=row_idx, column=3, value=status)
        cell_status.alignment = align_center
        if status == "Resolved":
            cell_status.fill = fill_low
            cell_status.font = font_low
        else:
            cell_status.fill = fill_action
            cell_status.font = font_action
        
        # Severity
        cell_sev = ws_details.cell(row=row_idx, column=4, value=severity)
        cell_sev.alignment = align_center
        if severity == "Critical":
            cell_sev.fill = fill_critical
            cell_sev.font = font_critical
        elif severity == "Medium":
            cell_sev.fill = fill_medium
            cell_sev.font = font_medium
        elif severity == "Low":
            cell_sev.fill = fill_low
            cell_sev.font = font_low
            
        # Improvement
        cell_imp = ws_details.cell(row=row_idx, column=5, value=improvement)
        cell_imp.alignment = align_wrap_left
        
        # Files
        cell_files = ws_details.cell(row=row_idx, column=6, value=files)
        cell_files.alignment = align_wrap_left
        
        # Zebra styling & thin borders
        for c in range(1, 7):
            cell = ws_details.cell(row=row_idx, column=c)
            cell.border = thin_border
            if cell.font.name != "Calibri" or cell.font.color is None:
                cell.font = font_regular
            if row_idx % 2 == 0 and c not in [3, 4]:
                cell.fill = fill_zebra
                
    # Style widths and heights
    ws_details.column_dimensions["A"].width = 8
    ws_details.column_dimensions["B"].width = 25
    ws_details.column_dimensions["C"].width = 16
    ws_details.column_dimensions["D"].width = 12
    ws_details.column_dimensions["E"].width = 75
    ws_details.column_dimensions["F"].width = 40
    
    for r in range(5, 15):
        ws_details.row_dimensions[r].height = 70
        
    wb.save(output_path)

if __name__ == "__main__":
    import sys
    out_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    target = os.path.join(out_dir, "production_readiness_report.xlsx")
    try:
        create_readiness_report(target)
        print(f"Excel report created at {target}")
    except PermissionError:
        fallback = os.path.join(out_dir, "production_readiness_report_updated.xlsx")
        print(f"Warning: {target} is locked. Saving fallback report to {fallback}")
        try:
            create_readiness_report(fallback)
            print(f"Excel report created at {fallback}")
        except Exception as e:
            print(f"Failed to generate fallback Excel report: {e}")
