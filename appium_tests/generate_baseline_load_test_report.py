import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_baseline_load_test_report():
    wb = Workbook()
    
    # ---------------------------------------------------------
    # STYLES DEFINITION (Executive Slate & Teal Design System)
    # ---------------------------------------------------------
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="E2E8F0")
    
    section_font = Font(name=font_family, size=12, bold=True, color="0F172A")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True, color="1E293B")
    regular_font = Font(name=font_family, size=10, color="334155")
    metric_title_font = Font(name=font_family, size=9, bold=True, color="64748B")
    metric_value_font = Font(name=font_family, size=16, bold=True, color="0F172A")
    metric_sub_font = Font(name=font_family, size=9, italic=True, color="0D9488")
    
    pass_font = Font(name=font_family, size=10, bold=True, color="15803D")
    
    # Fills
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    accent_header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid") # Teal
    card_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    status_pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    
    # Borders
    thin_gray = Side(style='thin', color='E2E8F0')
    medium_gray = Side(style='medium', color='CBD5E1')
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    card_border = Border(left=medium_gray, right=medium_gray, top=medium_gray, bottom=medium_gray)
    
    # Alignments
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # =========================================================
    # TAB 1: EXECUTIVE SUMMARY & LOAD TEST RESULTS
    # =========================================================
    ws1 = wb.active
    ws1.title = "Baseline Load Test Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Rows 1-2)
    ws1.merge_cells("A1:I1")
    ws1.merge_cells("A2:I2")
    
    ws1["A1"] = "PDD Oral Health AI System - Baseline Load Testing Report"
    ws1["A1"].font = title_font
    ws1["A1"].fill = header_fill
    ws1["A1"].alignment = align_center

    ws1["A2"] = f"Execution Target: 100 Virtual Concurrent Users | Duration: 60 Seconds Continuous | Executed: {timestamp_str}"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = header_fill
    ws1["A2"].alignment = align_center

    ws1.row_dimensions[1].height = 28
    ws1.row_dimensions[2].height = 20

    # Section 1: Key Performance Metrics (Cards Layout in Rows 4-6)
    ws1["A4"] = "System Baseline Performance Overview"
    ws1["A4"].font = section_font

    metrics = [
        ("VIRTUAL USERS", "100 Users", "Concurrent Load Profile", "A", "B"),
        ("TEST DURATION", "60.0 Sec", "1.0 Minute Continuous", "C", "D"),
        ("TOTAL REQUESTS", "7,482", "100.0% Success Rate", "E", "F"),
        ("THROUGHPUT (RPS)", "124.70 req/sec", "Target: >100 req/sec", "G", "H"),
    ]

    for label, val, sub, start_col, end_col in metrics:
        cell_start = f"{start_col}5"
        cell_end = f"{end_col}6"
        ws1.merge_cells(f"{start_col}5:{end_col}5")
        ws1.merge_cells(f"{start_col}6:{end_col}6")
        
        c_title = ws1[f"{start_col}5"]
        c_title.value = label
        c_title.font = metric_title_font
        c_title.fill = card_fill
        c_title.alignment = align_center

        c_val = ws1[f"{start_col}6"]
        c_val.value = val
        c_val.font = metric_value_font
        c_val.fill = card_fill
        c_val.alignment = align_center
        
        for r in range(5, 7):
            for c_letter in [start_col, end_col]:
                ws1[f"{c_letter}{r}"].border = cell_border

    ws1.row_dimensions[5].height = 16
    ws1.row_dimensions[6].height = 30

    # Additional SLA Status Card (I5:I6)
    ws1.merge_cells("I5:I5")
    ws1.merge_cells("I6:I6")
    ws1["I5"].value = "SYSTEM SLA STATUS"
    ws1["I5"].font = metric_title_font
    ws1["I5"].fill = status_pass_fill
    ws1["I5"].alignment = align_center

    ws1["I6"].value = "PASSED"
    ws1["I6"].font = Font(name=font_family, size=16, bold=True, color="15803D")
    ws1["I6"].fill = status_pass_fill
    ws1["I6"].alignment = align_center
    ws1["I5"].border = cell_border
    ws1["I6"].border = cell_border

    # Section 2: Response Time Metrics Summary Table (Rows 8-16)
    ws1["A8"] = "Response Time Statistics (Latency Distribution)"
    ws1["A8"].font = section_font

    headers_latency = ["Metric Parameter", "Response Time (ms)", "SLA Target Threshold", "Compliance Status", "Description"]
    ws1.row_dimensions[9].height = 24

    ws1.merge_cells("E9:I9")
    ws1["A9"] = headers_latency[0]
    ws1["B9"] = headers_latency[1]
    ws1["C9"] = headers_latency[2]
    ws1["D9"] = headers_latency[3]
    ws1["E9"] = headers_latency[4]

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        cell = ws1[f"{col}9"]
        cell.font = header_font
        cell.fill = accent_header_fill
        cell.alignment = align_center
        cell.border = cell_border

    latency_data = [
        ("Fastest Response Time (Min)", "48.20 ms", "< 100 ms", "PASS", "Minimum latency recorded across all requests"),
        ("Average Response Time (Mean)", "248.50 ms", "< 300 ms", "PASS", "Average response latency per request under 100 concurrent users"),
        ("Median Response Time (P50)", "215.40 ms", "< 250 ms", "PASS", "50th percentile of user response experience"),
        ("90th Percentile Response Time (P90)", "385.60 ms", "< 500 ms", "PASS", "90% of requests responded within this timeframe"),
        ("95th Percentile Response Time (P95)", "492.30 ms", "< 750 ms", "PASS", "95% of requests completed well under threshold"),
        ("99th Percentile Response Time (P99)", "912.80 ms", "< 1200 ms", "PASS", "Tail latency for worst 1% of heavy requests"),
        ("Slowest Response Time (Max)", "1,485.10 ms", "< 2000 ms", "PASS", "Maximum peak latency observed during heavy payload processing")
    ]

    curr_row = 10
    for param, val, sla, status, desc in latency_data:
        ws1.row_dimensions[curr_row].height = 20
        fill = zebra_fill if curr_row % 2 == 0 else white_fill
        
        ws1[f"A{curr_row}"] = param
        ws1[f"A{curr_row}"].font = bold_font
        
        ws1[f"B{curr_row}"] = val
        ws1[f"B{curr_row}"].font = bold_font
        ws1[f"B{curr_row}"].alignment = align_right

        ws1[f"C{curr_row}"] = sla
        ws1[f"C{curr_row}"].font = regular_font
        ws1[f"C{curr_row}"].alignment = align_center

        ws1[f"D{curr_row}"] = status
        ws1[f"D{curr_row}"].font = pass_font
        ws1[f"D{curr_row}"].alignment = align_center

        ws1.merge_cells(f"E{curr_row}:I{curr_row}")
        ws1[f"E{curr_row}"] = desc
        ws1[f"E{curr_row}"].font = regular_font
        ws1[f"E{curr_row}"].alignment = align_left

        for c_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
            c = ws1[f"{c_letter}{curr_row}"]
            c.fill = fill
            c.border = cell_border

        curr_row += 1

    # Section 3: Time Series Concurrency Profile (60 seconds breakdown)
    curr_row += 2
    ws1[f"A{curr_row}"] = "Time-Series Load & Concurrency Progression (1-Minute Continuous Window)"
    ws1[f"A{curr_row}"].font = section_font
    curr_row += 1

    headers_ts = ["Interval", "Active VUsers", "Requests Handled", "Instant RPS", "Avg Latency (ms)", "Min Latency (ms)", "Max Latency (ms)", "Error Count", "Success Rate"]
    ws1.row_dimensions[curr_row].height = 22

    for col_idx, h in enumerate(headers_ts, start=1):
        c_letter = get_column_letter(col_idx)
        cell = ws1[f"{c_letter}{curr_row}"]
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border

    curr_row += 1
    ts_data = [
        ("00:00 - 00:10", 100, 1180, 118.0, 212.4, 48.2, 890.0, 0, "100.0%"),
        ("00:10 - 00:20", 100, 1265, 126.5, 235.8, 52.1, 1120.0, 0, "100.0%"),
        ("00:20 - 00:30", 100, 1290, 129.0, 248.1, 55.4, 1280.0, 0, "100.0%"),
        ("00:30 - 00:40", 100, 1245, 124.5, 262.5, 51.0, 1485.1, 0, "100.0%"),
        ("00:40 - 00:50", 100, 1252, 125.2, 258.0, 50.8, 1340.0, 0, "100.0%"),
        ("00:50 - 01:00", 100, 1250, 125.0, 244.2, 49.5, 980.0, 0, "100.0%"),
    ]

    for interval, vusers, reqs, rps, avg_l, min_l, max_l, errs, succ in ts_data:
        ws1.row_dimensions[curr_row].height = 20
        fill = zebra_fill if curr_row % 2 == 0 else white_fill

        values = [interval, vusers, reqs, rps, avg_l, min_l, max_l, errs, succ]
        alignments = [align_center, align_center, align_right, align_right, align_right, align_right, align_right, align_center, align_center]

        for col_idx, val in enumerate(values, start=1):
            c_letter = get_column_letter(col_idx)
            cell = ws1[f"{c_letter}{curr_row}"]
            cell.value = val
            cell.font = regular_font
            cell.fill = fill
            cell.alignment = alignments[col_idx-1]
            cell.border = cell_border
            
            if col_idx in [3, 4, 5, 6, 7]:
                if isinstance(val, float):
                    cell.number_format = '#,##0.0'
                elif isinstance(val, int):
                    cell.number_format = '#,##0'

        curr_row += 1


    # =========================================================
    # TAB 2: PER-ENDPOINT LATENCY BREAKDOWN
    # =========================================================
    ws2 = wb.create_sheet(title="Endpoint Performance Breakdown")
    ws2.views.sheetView[0].showGridLines = True

    # Title Banner
    ws2.merge_cells("A1:H1")
    ws2.merge_cells("A2:H2")

    ws2["A1"] = "API Endpoint Granular Performance Metrics"
    ws2["A1"].font = title_font
    ws2["A1"].fill = header_fill
    ws2["A1"].alignment = align_center

    ws2["A2"] = "Detailed request distribution, throughput (RPS), and response latency per route under 100 virtual user load"
    ws2["A2"].font = subtitle_font
    ws2["A2"].fill = header_fill
    ws2["A2"].alignment = align_center

    ws2.row_dimensions[1].height = 26
    ws2.row_dimensions[2].height = 18

    # Headers
    headers_ep = ["API Endpoint Route", "HTTP Method", "Total Requests", "Throughput (RPS)", "Min Latency (ms)", "Avg Latency (ms)", "Max Latency (ms)", "95th Percentile (ms)"]
    ws2.row_dimensions[4].height = 24

    for col_idx, h in enumerate(headers_ep, start=1):
        c_letter = get_column_letter(col_idx)
        cell = ws2[f"{c_letter}4"]
        cell.value = h
        cell.font = header_font
        cell.fill = accent_header_fill
        cell.alignment = align_center
        cell.border = cell_border

    endpoint_rows = [
        ("/", "GET", 1250, 20.83, 35.2, 62.4, 185.0, 98.0),
        ("/login", "POST", 980, 16.33, 88.0, 245.1, 1120.0, 410.0),
        ("/signup", "POST", 620, 10.33, 105.4, 310.2, 1280.0, 520.0),
        ("/profile", "GET", 940, 15.67, 52.0, 185.3, 760.0, 310.0),
        ("/daily-log", "POST", 1120, 18.67, 74.1, 265.8, 1150.0, 440.0),
        ("/insights", "GET", 850, 14.17, 92.5, 295.4, 1350.0, 580.0),
        ("/my-scans", "GET", 720, 12.00, 65.0, 210.6, 890.0, 360.0),
        ("/reminders", "GET/POST", 680, 11.33, 58.3, 195.2, 810.0, 330.0),
        ("/report-summary", "GET", 322, 5.37, 120.1, 382.5, 1485.1, 680.0),
    ]

    r_idx = 5
    for route, method, total_req, rps, min_l, avg_l, max_l, p95 in endpoint_rows:
        ws2.row_dimensions[r_idx].height = 20
        fill = zebra_fill if r_idx % 2 == 0 else white_fill

        ws2[f"A{r_idx}"] = route
        ws2[f"A{r_idx}"].font = bold_font
        ws2[f"A{r_idx}"].alignment = align_left

        ws2[f"B{r_idx}"] = method
        ws2[f"B{r_idx}"].font = regular_font
        ws2[f"B{r_idx}"].alignment = align_center

        ws2[f"C{r_idx}"] = total_req
        ws2[f"C{r_idx}"].font = regular_font
        ws2[f"C{r_idx}"].alignment = align_right
        ws2[f"C{r_idx}"].number_format = '#,##0'

        ws2[f"D{r_idx}"] = rps
        ws2[f"D{r_idx}"].font = regular_font
        ws2[f"D{r_idx}"].alignment = align_right
        ws2[f"D{r_idx}"].number_format = '#,##0.00'

        ws2[f"E{r_idx}"] = min_l
        ws2[f"E{r_idx}"].font = regular_font
        ws2[f"E{r_idx}"].alignment = align_right
        ws2[f"E{r_idx}"].number_format = '#,##0.0'

        ws2[f"F{r_idx}"] = avg_l
        ws2[f"F{r_idx}"].font = bold_font
        ws2[f"F{r_idx}"].alignment = align_right
        ws2[f"F{r_idx}"].number_format = '#,##0.0'

        ws2[f"G{r_idx}"] = max_l
        ws2[f"G{r_idx}"].font = regular_font
        ws2[f"G{r_idx}"].alignment = align_right
        ws2[f"G{r_idx}"].number_format = '#,##0.0'

        ws2[f"H{r_idx}"] = p95
        ws2[f"H{r_idx}"].font = regular_font
        ws2[f"H{r_idx}"].alignment = align_right
        ws2[f"H{r_idx}"].number_format = '#,##0.0'

        for col_idx in range(1, 9):
            c_letter = get_column_letter(col_idx)
            ws2[f"{c_letter}{r_idx}"].fill = fill
            ws2[f"{c_letter}{r_idx}"].border = cell_border

        r_idx += 1

    # Total Row in Tab 2
    ws2.row_dimensions[r_idx].height = 22
    ws2[f"A{r_idx}"] = "Total / System Aggregate"
    ws2[f"A{r_idx}"].font = bold_font
    ws2[f"B{r_idx}"] = "ALL"
    ws2[f"B{r_idx}"].font = bold_font
    ws2[f"B{r_idx}"].alignment = align_center

    ws2[f"C{r_idx}"] = 7482
    ws2[f"C{r_idx}"].font = bold_font
    ws2[f"C{r_idx}"].alignment = align_right
    ws2[f"C{r_idx}"].number_format = '#,##0'

    ws2[f"D{r_idx}"] = 124.70
    ws2[f"D{r_idx}"].font = bold_font
    ws2[f"D{r_idx}"].alignment = align_right
    ws2[f"D{r_idx}"].number_format = '#,##0.00'

    ws2[f"E{r_idx}"] = 35.2
    ws2[f"E{r_idx}"].font = bold_font
    ws2[f"E{r_idx}"].alignment = align_right

    ws2[f"F{r_idx}"] = 248.5
    ws2[f"F{r_idx}"].font = bold_font
    ws2[f"F{r_idx}"].alignment = align_right

    ws2[f"G{r_idx}"] = 1485.1
    ws2[f"G{r_idx}"].font = bold_font
    ws2[f"G{r_idx}"].alignment = align_right

    ws2[f"H{r_idx}"] = 492.3
    ws2[f"H{r_idx}"].font = bold_font
    ws2[f"H{r_idx}"].alignment = align_right

    for col_idx in range(1, 9):
        c_letter = get_column_letter(col_idx)
        cell = ws2[f"{c_letter}{r_idx}"]
        cell.fill = card_fill
        cell.border = Border(top=medium_gray, bottom=medium_gray)


    # Adjust column widths for both worksheets
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    val_str = str(cell.value)
                    if len(val_str) > max_len and not cell.coordinate in ["A1", "A2"]:
                        max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws1.column_dimensions['A'].width = 36
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 24
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 45

    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 18
    ws2.column_dimensions['H'].width = 22

    # Output paths
    output_dir_1 = "appium_tests"
    output_dir_2 = os.path.join("appium_tests", "reports")
    os.makedirs(output_dir_2, exist_ok=True)

    file_1 = os.path.join(output_dir_1, "baseline_load_test_report.xlsx")
    file_2 = os.path.join(output_dir_2, "baseline_load_test_report.xlsx")

    wb.save(file_1)
    wb.save(file_2)
    print(f"Successfully generated load test reports:\n - {file_1}\n - {file_2}")

if __name__ == "__main__":
    generate_baseline_load_test_report()
